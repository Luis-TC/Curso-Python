from openpyxl import load_workbook

# Ruta del archivo Excel
file_path = "C:/Users/d_gne/Documents/CURSO PYTHON/Curso-Python/Clase 10/archivo_excel/archivo.xlsx"

# Abrir el archivo Excel
workbook = load_workbook(filename=file_path)

# Seleccionar una hoja específica
sheet = workbook.active  # O workbook['nombre_hoja']

# Leer el contenido de las celdas
for row in sheet.iter_rows(values_only=True):
    print(row)